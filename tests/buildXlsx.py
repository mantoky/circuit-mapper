#!/usr/bin/env python3
"""
Realiza o workbook-spec.json em .xlsx usando openpyxl.
Espelha exatamente a logica de src/export/xlsxExport.js (exceljs) usada no app,
servindo como verificacao independente do spec.
"""
import json, os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, 'output')
spec = json.load(open(os.path.join(OUT, 'workbook-spec.json'), encoding='utf-8'))
S = spec['style']

def rgb(argb): return argb[2:] if len(argb) == 8 else argb
def fill(argb): return PatternFill('solid', fgColor=rgb(argb))
thin = Side(style='thin', color=rgb(S['border']))
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook(); wb.remove(wb.active)
wb.properties.creator = spec['meta']['creator']
wb.properties.title   = spec['meta']['title']

for sheet in spec['sheets']:
    ws = wb.create_sheet(sheet['name'][:31])
    for i, w in enumerate(sheet.get('widths') or [], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    header_row_idx = None
    for r in sheet['rows']:
        cells = r.get('cells') or []
        spans = r.get('spans') or []
        key_cols = set()

        if spans:
            # distribui as celulas conforme os spans (mesma logica do exceljs)
            flat, ranges, col = {}, [], 1
            for i, v in enumerate(cells):
                span = spans[i] if i < len(spans) else 1
                flat[col] = v
                if i % 2 == 0: key_cols.add(col)
                if span > 1: ranges.append((col, col + span - 1))
                col += span
            width = len(sheet['widths'])
            ws.append([flat.get(c) for c in range(1, max(width, col - 1) + 1)])
            row = ws.max_row
            for a, b in ranges:
                try: ws.merge_cells(start_row=row, start_column=a, end_row=row, end_column=b)
                except Exception: pass
        else:
            ws.append(cells)
            row = ws.max_row
            if r.get('merge') and cells:
                try: ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=r['merge'])
                except Exception: pass

        style = r.get('style')
        status = r.get('status')
        ncols = max(len(cells), 1)

        for c in range(1, len(sheet['widths']) + 1):
            cell = ws.cell(row=row, column=c)
            if style == 'title':
                ws.row_dimensions[row].height = 26
                cell.fill = fill(S['charcoal']); cell.font = Font(name='Roboto', size=14, bold=True, color=rgb(S['yellow']))
                cell.alignment = Alignment(vertical='center', horizontal='left', indent=1)
            elif style == 'subtitle':
                cell.fill = fill(S['charcoalSoft']); cell.font = Font(name='Roboto', size=10, italic=True, color=rgb(S['textLight']))
            elif style == 'section':
                cell.fill = fill(S['yellow']); cell.font = Font(name='Roboto', size=10, bold=True, color=rgb(S['textDark']))
            elif style == 'header':
                header_row_idx = row
                ws.row_dimensions[row].height = 32
                cell.fill = fill(S['charcoal']); cell.font = Font(name='Roboto', size=8.5, bold=True, color=rgb(S['yellow']))
                cell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True); cell.border = BORDER
            elif style in ('kv', 'kvline'):
                in_range = (c in key_cols or ws.cell(row=row, column=c).value is not None) if spans else (c <= ncols)
                if in_range:
                    is_key = (c in key_cols) if spans else ((c == 1) if style == 'kv' else (c % 2 == 1))
                    cell.font = Font(name='Roboto', size=9, bold=is_key)
                    if is_key: cell.fill = fill(S['gray'])
                    cell.border = BORDER
            elif style == 'total':
                cell.fill = fill(S['yellow']); cell.font = Font(name='Roboto', size=9, bold=True, color=rgb(S['textDark'])); cell.border = BORDER
            elif style == 'verdict':
                if c <= ncols:
                    cell.fill = fill(S['charcoal']); cell.font = Font(name='Roboto', size=11, bold=True, color=rgb(S['yellow'])); cell.border = BORDER
            elif style == 'note':
                if c <= ncols:
                    cell.font = Font(name='Roboto', size=8, italic=True, color='666666')
            else:
                cell.font = Font(name='Roboto', size=9)
                cell.border = BORDER
                cell.alignment = Alignment(vertical='center')
                bg = {'error': S['err'], 'warn': S['warn']}.get(status)
                if bg: cell.fill = fill(bg)

    if sheet.get('freeze'): ws.freeze_panes = sheet['freeze']
    if header_row_idx and ws.max_row > header_row_idx:
        ws.auto_filter.ref = (f"A{header_row_idx}:"
                             f"{get_column_letter(len(sheet['widths']))}{ws.max_row}")
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

base = json.load(open(os.path.join(OUT, 'metrics.json'), encoding='utf-8'))['arquivoBase']
dest = os.path.join(OUT, f'{base}_QUADRO-DE-CARGAS.xlsx')
wb.save(dest)
print(f'XLSX gerado: {os.path.basename(dest)}  |  {len(wb.sheetnames)} abas: {", ".join(wb.sheetnames)}')
